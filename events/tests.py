import csv
import json
import os
import tempfile
from datetime import timedelta
from io import StringIO

from django.contrib.auth import get_user_model
from django.core.management import call_command
from django.test import TestCase
from django.utils import timezone
from openpyxl import Workbook
from rest_framework.test import APIClient

from .models import Category, City, Event, Favorite, Tag, Venue


class EventsApiTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.city = City.objects.create(name="Apeldoorn", slug="apeldoorn")
        self.category, _ = Category.objects.get_or_create(
            slug="muziek",
            defaults={
                "name": "Muziek",
                "kind": Event.Kind.EVENT,
            },
        )
        self.venue = Venue.objects.create(
            name="Gigant",
            city=self.city,
            address="Nieuwstraat 1, Apeldoorn",
            latitude="52.211157",
            longitude="5.965546",
        )
        self.audience_tag, _ = Tag.objects.get_or_create(
            slug="vrienden",
            defaults={
                "name": "Vrienden",
                "facet": Tag.Facet.AUDIENCE,
            },
        )
        self.vibe_tag, _ = Tag.objects.get_or_create(
            slug="cultureel",
            defaults={
                "name": "Cultureel",
                "facet": Tag.Facet.VIBE,
            },
        )

        self.event = Event.objects.create(
            title="Live muziek in Gigant",
            slug="live-muziek-in-gigant-apeldoorn",
            city=self.city,
            venue=self.venue,
            category=self.category,
            summary="Een sterke avondmatch.",
            description="Volledige beschrijving",
            image_url="https://example.com/live-muziek.jpg",
            start_at=timezone.now() + timedelta(hours=2),
            is_free=False,
            price_min="18.00",
            date_text="Vanavond",
            raw_date_text="Vanavond",
            kind=Event.Kind.EVENT,
            is_featured=True,
        )
        self.active_event = Event.objects.create(
            title="Nu bezig in CODA",
            slug="nu-bezig-in-coda-apeldoorn",
            city=self.city,
            venue=self.venue,
            category=self.category,
            start_at=timezone.now() - timedelta(hours=1),
            end_at=timezone.now() + timedelta(hours=1),
            date_text="Vandaag",
            raw_date_text="Vandaag",
            kind=Event.Kind.EVENT,
        )
        self.unknown_date_event = Event.objects.create(
            title="Datums volgen nog",
            slug="datums-volgen-nog-apeldoorn",
            city=self.city,
            date_text="Datum volgt",
            raw_date_text="Datum volgt",
            kind=Event.Kind.EVENT,
        )
        self.missing_date_event = Event.objects.create(
            title="Nog zonder datum",
            slug="nog-zonder-datum-apeldoorn",
            city=self.city,
            kind=Event.Kind.EVENT,
        )
        self.event.tags.add(self.audience_tag, self.vibe_tag)
        self.active_event.tags.add(self.audience_tag)

    def test_events_list_backwards_compatible_and_extended(self):
        response = self.client.get("/api/events/?city=apeldoorn&q=Live muziek")
        self.assertEqual(response.status_code, 200)

        payload = response.json()
        self.assertIn("count", payload)
        self.assertIn("results", payload)
        self.assertEqual(payload["count"], 1)

        result = payload["results"][0]
        self.assertEqual(result["id"], self.event.id)
        self.assertEqual(result["title"], self.event.title)
        self.assertEqual(result["city"], "apeldoorn")
        self.assertEqual(result["venue"], "Gigant")
        self.assertIn("summary", result)
        self.assertIn("slug", result)
        self.assertIn("category", result)
        self.assertIn("audiences", result)

    def test_event_detail_by_id_and_slug(self):
        response = self.client.get(f"/api/events/{self.event.id}/")
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["id"], self.event.id)

        slug_response = self.client.get(f"/api/events/by-slug/{self.event.slug}/")
        self.assertEqual(slug_response.status_code, 200)
        self.assertEqual(slug_response.json()["slug"], self.event.slug)

    def test_metadata_endpoints(self):
        self.assertEqual(self.client.get("/api/cities/").status_code, 200)
        self.assertEqual(self.client.get("/api/cities/apeldoorn/").status_code, 200)
        self.assertEqual(self.client.get("/api/categories/").status_code, 200)
        self.assertEqual(self.client.get("/api/tags/").status_code, 200)

    def test_new_filters(self):
        response = self.client.get(
            "/api/events/?city=apeldoorn&category=muziek&audience=vrienden&vibe=cultureel&featured=1"
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["count"], 1)

    def test_ongoing_flag_and_filter_use_real_time_window(self):
        active_detail = self.client.get(f"/api/events/{self.active_event.id}/")
        self.assertEqual(active_detail.status_code, 200)
        self.assertTrue(active_detail.json()["is_ongoing"])
        self.assertEqual(active_detail.json()["status"], "Nu bezig")

        unknown_date_detail = self.client.get(f"/api/events/{self.unknown_date_event.id}/")
        self.assertEqual(unknown_date_detail.status_code, 200)
        self.assertFalse(unknown_date_detail.json()["is_ongoing"])
        self.assertEqual(unknown_date_detail.json()["status"], "Datum volgt")

        missing_date_detail = self.client.get(f"/api/events/{self.missing_date_event.id}/")
        self.assertEqual(missing_date_detail.status_code, 200)
        self.assertFalse(missing_date_detail.json()["is_ongoing"])
        self.assertEqual(missing_date_detail.json()["status"], "Datum nog te controleren")

        ongoing_response = self.client.get("/api/events/?city=apeldoorn&ongoing=1")
        self.assertEqual(ongoing_response.status_code, 200)
        self.assertEqual(ongoing_response.json()["count"], 1)
        self.assertEqual(ongoing_response.json()["results"][0]["id"], self.active_event.id)

        not_ongoing_response = self.client.get("/api/events/?city=apeldoorn&ongoing=0")
        self.assertEqual(not_ongoing_response.status_code, 200)
        returned_ids = {item["id"] for item in not_ongoing_response.json()["results"]}
        self.assertIn(self.event.id, returned_ids)
        self.assertIn(self.unknown_date_event.id, returned_ids)
        self.assertIn(self.missing_date_event.id, returned_ids)
        self.assertNotIn(self.active_event.id, returned_ids)

    def test_events_data_quality_endpoint_returns_counts_and_incomplete_records(self):
        response = self.client.get("/api/data-quality/events/")
        self.assertEqual(response.status_code, 200)

        payload = response.json()
        self.assertEqual(payload["total_events"], 4)
        self.assertEqual(payload["events_per_city"], {"apeldoorn": 4})
        self.assertEqual(payload["with_category"], 2)
        self.assertEqual(payload["without_category"], 2)
        self.assertEqual(payload["with_summary"], 1)
        self.assertEqual(payload["without_summary"], 3)
        self.assertEqual(payload["with_image_url"], 1)
        self.assertEqual(payload["without_image_url"], 3)
        self.assertEqual(payload["with_coordinates"], 2)
        self.assertEqual(payload["without_coordinates"], 2)
        self.assertEqual(payload["with_venue"], 2)
        self.assertEqual(payload["without_venue"], 2)
        self.assertEqual(payload["with_start_at"], 2)
        self.assertEqual(payload["without_start_at"], 2)
        self.assertEqual(payload["with_tags"], 2)
        self.assertEqual(payload["without_tags"], 2)
        self.assertEqual(payload["featured"], 1)
        self.assertEqual(payload["hidden_gem"], 0)
        self.assertEqual(payload["discover_ready"], 1)
        self.assertEqual(payload["not_discover_ready"], 3)

        incomplete_by_id = {item["id"]: item for item in payload["incomplete_records"]}
        self.assertEqual(len(incomplete_by_id), 3)
        self.assertFalse(incomplete_by_id[self.active_event.id]["discover_ready"])
        self.assertIn("summary", incomplete_by_id[self.active_event.id]["missing_fields"])
        self.assertIn("image_url", incomplete_by_id[self.unknown_date_event.id]["missing_fields"])

    def test_events_data_quality_endpoint_supports_city_filter(self):
        other_city = City.objects.create(name="Deventer", slug="deventer")
        Event.objects.create(
            title="Deventer compleet",
            city=other_city,
            category=self.category,
            summary="Compleet genoeg voor ontdek.",
            image_url="https://example.com/deventer.jpg",
            address="Brink 1, Deventer",
            date_text="Vandaag",
            is_hidden_gem=True,
        )

        response = self.client.get("/api/data-quality/events/?city=deventer")
        self.assertEqual(response.status_code, 200)

        payload = response.json()
        self.assertEqual(payload["total_events"], 1)
        self.assertEqual(payload["events_per_city"], {"deventer": 1})
        self.assertEqual(payload["discover_ready"], 1)
        self.assertEqual(payload["not_discover_ready"], 0)
        self.assertEqual(payload["hidden_gem"], 1)
        self.assertEqual(payload["incomplete_records"], [])


class CityContentApiTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.city = City.objects.create(name="Harderwijk", slug="harderwijk")
        self.lelystad = City.objects.create(name="Lelystad", slug="lelystad")
        self.other_city = City.objects.create(name="Deventer", slug="deventer")
        self.activity_category, _ = Category.objects.get_or_create(
            slug="activiteit",
            defaults={
                "name": "Activiteit",
                "kind": Event.Kind.ACTIVITY,
            },
        )
        self.food_category, _ = Category.objects.get_or_create(
            slug="restaurant",
            defaults={
                "name": "Restaurant",
                "kind": Event.Kind.FOOD_DRINK,
            },
        )
        self.venue = Venue.objects.create(
            name="Museum Harderwijk",
            slug="museum-harderwijk",
            city=self.city,
            address="Markt 1",
            latitude="52.350000",
            longitude="5.620000",
        )
        self.tag = Tag.objects.create(
            name="Gezinnen",
            slug="gezinnen",
            facet=Tag.Facet.AUDIENCE,
        )
        self.outing = Event.objects.create(
            title="Museumwandeling Harderwijk",
            slug="museumwandeling-harderwijk",
            kind=Event.Kind.ACTIVITY,
            city=self.city,
            venue=self.venue,
            category=self.activity_category,
            summary="Een wandeling langs het museum.",
            description="Route door de stad met museumstop.",
            source_url="https://example.com/museumwandeling",
            ticket_url="nan",
            image_url="NaN",
            price_note="",
        )
        self.outing.tags.add(self.tag)
        self.food = Event.objects.create(
            title="Lunch aan de haven",
            slug="lunch-aan-de-haven",
            kind=Event.Kind.FOOD_DRINK,
            city=self.city,
            category=self.food_category,
            summary="Lunchplek aan de haven.",
            description="Restaurant met terras.",
            address="Havenkade 2",
            latitude="52.351000",
            longitude="5.621000",
            source_url="https://example.com/lunch-aan-de-haven",
            price_note="n/a",
        )
        Event.objects.create(
            title="Deventer wandeling",
            slug="deventer-wandeling",
            kind=Event.Kind.ACTIVITY,
            city=self.other_city,
            category=self.activity_category,
            source_url="https://example.com/deventer-wandeling",
        )
        self.lelystad_outing = Event.objects.create(
            title="Lelystad stadswandeling",
            slug="lelystad-stadswandeling",
            kind=Event.Kind.ACTIVITY,
            city=self.lelystad,
            category=self.activity_category,
            source="city_content:outings",
            source_url="https://example.com/lelystad-stadswandeling",
        )
        self.lelystad_food = Event.objects.create(
            title="Lelystad restaurant",
            slug="lelystad-restaurant",
            kind=Event.Kind.FOOD_DRINK,
            city=self.lelystad,
            category=self.food_category,
            source="city_content:food_drink",
            source_url="https://example.com/lelystad-restaurant",
        )
        for title, slug, kind in [
            ("Lelystad snackbar", "lelystad-snackbar", "snackbar"),
            ("Lelystad ijssalon", "lelystad-ijssalon", "ice_cream"),
            ("Lelystad strandpaviljoen", "lelystad-strandpaviljoen", "beach_pavilion"),
        ]:
            Event.objects.create(
                title=title,
                slug=slug,
                kind=kind,
                city=self.lelystad,
                category=self.food_category,
                source="city_content:food_drink",
                source_url=f"https://example.com/{slug}",
            )
        Event.objects.create(
            title="Lelystad food paviljoen",
            slug="lelystad-food-paviljoen",
            kind=Event.Kind.PLACE,
            city=self.lelystad,
            category=self.food_category,
            source="city_content:food_drink",
            source_url="https://example.com/lelystad-food-paviljoen",
        )
        Event.objects.create(
            title="Lelystad gewoon park",
            slug="lelystad-gewoon-park",
            kind=Event.Kind.PLACE,
            city=self.lelystad,
            category=self.activity_category,
            source="city_content:outings",
            source_url="https://example.com/lelystad-gewoon-park",
        )

    def test_city_content_endpoint_returns_200(self):
        response = self.client.get("/api/city-content/")
        self.assertEqual(response.status_code, 200)
        self.assertIn("results", response.json())

    def test_city_filter_returns_only_requested_city(self):
        response = self.client.get("/api/city-content/?city=harderwijk")
        self.assertEqual(response.status_code, 200)

        payload = response.json()
        self.assertEqual(payload["count"], 2)
        self.assertEqual({item["city"] for item in payload["results"]}, {"harderwijk"})

    def test_type_food_drink_returns_food_drink(self):
        response = self.client.get("/api/city-content/?city=harderwijk&type=food_drink")
        self.assertEqual(response.status_code, 200)

        payload = response.json()
        self.assertEqual(payload["count"], 1)
        self.assertEqual(payload["results"][0]["kind"], Event.Kind.FOOD_DRINK)

    def test_type_outings_excludes_food_drink(self):
        response = self.client.get("/api/city-content/?city=harderwijk&type=outings")
        self.assertEqual(response.status_code, 200)

        payload = response.json()
        self.assertEqual(payload["count"], 1)
        returned_kinds = {item["kind"] for item in payload["results"]}
        self.assertNotIn(Event.Kind.FOOD_DRINK, returned_kinds)

    def test_type_food_drink_includes_food_drink_subtypes(self):
        response = self.client.get("/api/city-content/?city=lelystad&type=food_drink&limit=20")
        self.assertEqual(response.status_code, 200)

        payload = response.json()
        self.assertEqual(payload["count"], 5)
        returned_kinds = {item["kind"] for item in payload["results"]}
        self.assertEqual(
            returned_kinds,
            {
                Event.Kind.FOOD_DRINK,
                "snackbar",
                "ice_cream",
                "beach_pavilion",
                Event.Kind.PLACE,
            },
        )

    def test_type_outings_excludes_food_drink_subtypes(self):
        response = self.client.get("/api/city-content/?city=lelystad&type=outings&limit=20")
        self.assertEqual(response.status_code, 200)

        payload = response.json()
        self.assertEqual(payload["count"], 2)
        returned_slugs = {item["slug"] for item in payload["results"]}
        self.assertEqual(returned_slugs, {"lelystad-stadswandeling", "lelystad-gewoon-park"})

    def test_harderwijk_counts_remain_unchanged_with_food_drink_mapping(self):
        food_response = self.client.get("/api/city-content/?city=harderwijk&type=food_drink")
        outings_response = self.client.get("/api/city-content/?city=harderwijk&type=outings")

        self.assertEqual(food_response.status_code, 200)
        self.assertEqual(outings_response.status_code, 200)
        self.assertEqual(food_response.json()["count"], 1)
        self.assertEqual(outings_response.json()["count"], 1)

    def test_query_filter_searches_without_crashing(self):
        response = self.client.get("/api/city-content/?city=harderwijk&query=museum")
        self.assertEqual(response.status_code, 200)

        payload = response.json()
        self.assertEqual(payload["count"], 1)
        self.assertEqual(payload["results"][0]["slug"], self.outing.slug)

    def test_nan_like_values_are_serialized_as_null(self):
        response = self.client.get("/api/city-content/?city=harderwijk&type=outings")
        self.assertEqual(response.status_code, 200)

        item = response.json()["results"][0]
        self.assertIsNone(item["ticket_url"])
        self.assertIsNone(item["image_url"])
        self.assertIsNone(item["price_note"])
        self.assertNotIn('"nan"', json.dumps(response.json()).lower())


class FavoritesCompatibilityTests(TestCase):
    def setUp(self):
        self.client = APIClient()
        self.user = get_user_model().objects.create_user(
            username="tester",
            password="secret123",
        )
        self.city = City.objects.create(name="Deventer", slug="deventer")
        self.event = Event.objects.create(
            title="Food Festival Deventer",
            slug="food-festival-deventer",
            city=self.city,
            kind=Event.Kind.FESTIVAL,
        )
        Favorite.objects.create(user=self.user, event=self.event)

    def test_favorites_endpoints_still_work(self):
        self.client.force_authenticate(self.user)

        list_response = self.client.get("/api/favorites/")
        self.assertEqual(list_response.status_code, 200)
        self.assertEqual(list_response.json()[0]["event_id"], self.event.id)

        events_response = self.client.get("/api/favorites/events/")
        self.assertEqual(events_response.status_code, 200)
        self.assertEqual(events_response.json()[0]["id"], self.event.id)


class ImportCommandTests(TestCase):
    def setUp(self):
        self.city = City.objects.create(name="Zwolle", slug="zwolle")

    def _create_excel_file(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append([
            "Naam/Activiteit",
            "Type",
            "Datum",
            "Omschrijving",
            "Website",
            "Category",
            "AudienceTags",
        ])
        sheet.append([
            "Zwolle Test Event",
            "Nieuw",
            "1 januari 2026",
            "Test omschrijving",
            "https://example.com/zwolle-test-event",
            "Cultuur",
            "Solo,Vrienden",
        ])

        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        temp_file.close()
        workbook.save(temp_file.name)
        return temp_file.name

    def _create_dry_run_excel_file(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append([
            "Naam/Activiteit",
            "Kind",
            "Datum",
            "Omschrijving",
            "Website",
            "SourceId",
            "Category",
            "Latitude",
            "Longitude",
            "IndoorOutdoor",
            "WeatherSuitability",
            "Gratis",
        ])
        sheet.append([
            "Bestaand Zwolle Event",
            "event",
            "1 januari 2026",
            "Zou update zijn",
            "https://example.com/existing-zwolle-event",
            "existing-1",
            "Cultuur",
            "52.512",
            "6.094",
            "indoor",
            "rain",
            "ja",
        ])
        sheet.append([
            "Nieuw Zwolle Event",
            "festival",
            "2 januari 2026",
            "Zou create zijn",
            "https://example.com/new-zwolle-event",
            "new-1",
            "Festival",
            "52.513",
            "6.095",
            "outdoor",
            "sun",
            "nee",
        ])
        sheet.append([
            "",
            "event",
            "3 januari 2026",
            "Mist titel",
            "https://example.com/missing-title",
            "missing-title",
            "Cultuur",
            "52.514",
            "6.096",
            "indoor",
            "all",
            "true",
        ])
        sheet.append([
            "Ongeldige waarden",
            "mystery",
            "32 januari 2026",
            "Heeft warnings",
            "https://example.com/invalid-values",
            "invalid-1",
            "Cultuur",
            "noord",
            "oost",
            "inside",
            "storm",
            "misschien",
        ])
        sheet.append([
            "Duplicate source",
            "event",
            "4 januari 2026",
            "Dubbele SourceId",
            "https://example.com/duplicate-source",
            "new-1",
            "Cultuur",
            "52.515",
            "6.097",
            "indoor",
            "all",
            "yes",
        ])

        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        temp_file.close()
        workbook.save(temp_file.name)
        return temp_file.name

    def test_import_command_is_idempotent(self):
        path = self._create_excel_file()
        try:
            call_command(
                "import_events_excel",
                file=path,
                city="zwolle",
                source="excel_zwolle_test",
            )
            first_event = Event.objects.get(source_url="https://example.com/zwolle-test-event")
            first_id = first_event.id

            call_command(
                "import_events_excel",
                file=path,
                city="zwolle",
                source="excel_zwolle_test",
            )

            self.assertEqual(
                Event.objects.filter(source_url="https://example.com/zwolle-test-event").count(),
                1,
            )
            self.assertEqual(
                Event.objects.get(source_url="https://example.com/zwolle-test-event").id,
                first_id,
            )
            self.assertEqual(Event.objects.get(id=first_id).raw_date_text, "1 januari 2026")
        finally:
            if os.path.exists(path):
                os.unlink(path)


    def test_import_command_dry_run_does_not_change_database(self):
        Event.objects.create(
            title="Bestaand Zwolle Event",
            city=self.city,
            source_url="https://example.com/existing-zwolle-event",
            source="excel_zwolle_test",
        )
        before_counts = {
            "events": Event.objects.count(),
            "categories": Category.objects.count(),
            "tags": Tag.objects.count(),
            "venues": Venue.objects.count(),
        }
        path = self._create_dry_run_excel_file()
        stdout = StringIO()
        stderr = StringIO()

        try:
            call_command(
                "import_events_excel",
                file=path,
                city="zwolle",
                source="excel_zwolle_test",
                dry_run=True,
                stdout=stdout,
                stderr=stderr,
            )
        finally:
            if os.path.exists(path):
                os.unlink(path)

        self.assertEqual(Event.objects.count(), before_counts["events"])
        self.assertEqual(Category.objects.count(), before_counts["categories"])
        self.assertEqual(Tag.objects.count(), before_counts["tags"])
        self.assertEqual(Venue.objects.count(), before_counts["venues"])

        output = stdout.getvalue()
        self.assertIn("DRY RUN: no database changes were written.", output)
        self.assertIn("rows=5", output)
        self.assertIn("would_create=2", output)
        self.assertIn("would_update=1", output)
        self.assertIn("would_skip=2", output)
        self.assertIn("errors=0", output)
        self.assertIn("missing required field title", output)
        self.assertIn("invalid date '32 januari 2026'", output)
        self.assertIn("invalid latitude 'noord'", output)
        self.assertIn("invalid longitude 'oost'", output)
        self.assertIn("unknown Kind value 'mystery'", output)
        self.assertIn("unknown IndoorOutdoor value 'inside'", output)
        self.assertIn("unknown WeatherSuitability value 'storm'", output)
        self.assertIn("unknown boolean value 'misschien'", output)
        self.assertIn("duplicate SourceId or dedupe_key 'new-1'", output)

    def test_import_command_dry_run_writes_report_without_database_changes(self):
        existing_event = Event.objects.create(
            title="Bestaand Zwolle Event",
            city=self.city,
            source_url="https://example.com/existing-zwolle-event",
            source="excel_zwolle_test",
        )
        path = self._create_dry_run_excel_file()
        report_dir = tempfile.mkdtemp()
        report_path = os.path.join(report_dir, "nested", "dry_run_report.csv")
        before_event_count = Event.objects.count()

        try:
            call_command(
                "import_events_excel",
                file=path,
                city="zwolle",
                source="excel_zwolle_test",
                dry_run=True,
                report_file=report_path,
            )

            self.assertEqual(Event.objects.count(), before_event_count)
            self.assertTrue(os.path.exists(report_path))

            with open(report_path, newline="", encoding="utf-8") as report_file:
                rows = list(csv.DictReader(report_file))

            self.assertEqual(len(rows), 5)
            self.assertEqual(rows[0]["action"], "would_update")
            self.assertEqual(rows[0]["event_id"], str(existing_event.id))
            self.assertEqual(rows[0]["discover_ready"], "false")
            self.assertIn("summary", rows[0]["missing_recommended_fields"])
            self.assertEqual(rows[1]["action"], "would_create")
            self.assertEqual(rows[1]["source_id"], "new-1")
            self.assertEqual(rows[2]["action"], "would_skip")
            self.assertIn("missing required field title", rows[2]["warnings"])
            self.assertEqual(rows[3]["action"], "would_create")
            self.assertIn("unknown Kind value 'mystery'", rows[3]["warnings"])
            self.assertIn("invalid latitude 'noord'", rows[3]["warnings"])
            self.assertEqual(rows[4]["action"], "would_skip")
            self.assertIn("duplicate SourceId or dedupe_key 'new-1'", rows[4]["warnings"])
        finally:
            if os.path.exists(path):
                os.unlink(path)

    def test_import_command_normal_import_writes_report_with_real_actions(self):
        path = self._create_excel_file()
        report_dir = tempfile.mkdtemp()
        report_path = os.path.join(report_dir, "normal_report.csv")

        try:
            call_command(
                "import_events_excel",
                file=path,
                city="zwolle",
                source="excel_zwolle_test",
                report_file=report_path,
            )

            with open(report_path, newline="", encoding="utf-8") as report_file:
                first_rows = list(csv.DictReader(report_file))

            self.assertEqual(len(first_rows), 1)
            self.assertEqual(first_rows[0]["action"], "created")
            self.assertTrue(first_rows[0]["event_id"])
            self.assertEqual(first_rows[0]["title"], "Zwolle Test Event")

            call_command(
                "import_events_excel",
                file=path,
                city="zwolle",
                source="excel_zwolle_test",
                report_file=report_path,
            )

            with open(report_path, newline="", encoding="utf-8") as report_file:
                second_rows = list(csv.DictReader(report_file))

            self.assertEqual(len(second_rows), 1)
            self.assertEqual(second_rows[0]["action"], "updated")
            self.assertEqual(second_rows[0]["event_id"], first_rows[0]["event_id"])
        finally:
            if os.path.exists(path):
                os.unlink(path)


class CityContentImportCommandTests(TestCase):
    def setUp(self):
        self.city = City.objects.create(name="Harderwijk", slug="harderwijk")
        self.category, _ = Category.objects.get_or_create(
            slug="cultuur",
            defaults={"name": "Cultuur", "kind": Event.Kind.EVENT},
        )

    def _create_city_content_excel_file(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Master template"
        sheet.append([
            "City",
            "Naam/Activiteit",
            "Slug",
            "Kind",
            "Category",
            "DatumRaw",
            "Venue",
            "Address",
            "Latitude",
            "Longitude",
            "Summary",
            "Omschrijving",
            "SourceUrl",
            "ImageUrl",
            "PriceNote",
            "IndoorOutdoor",
            "WeatherSuitability",
            "LastCheckedAt",
        ])
        sheet.append([
            "harderwijk",
            "Nieuwe stadswandeling",
            "nieuwe-stadswandeling",
            "activity",
            "Cultuur",
            "1 januari 2026",
            "Binnenstad",
            "Markt 1",
            "52.350000",
            "5.620000",
            "Een compacte stadswandeling.",
            "Een volledige beschrijving van de stadswandeling door Harderwijk.",
            "https://example.com/nieuwe-stadswandeling",
            "",
            "",
            "outdoor",
            "sun",
            "2026-05-13",
        ])
        sheet.append([
            "harderwijk",
            "Bestaande stadswandeling",
            "bestaande-stadswandeling",
            "activity",
            "Cultuur",
            "2 januari 2026",
            "Binnenstad",
            "Markt 2",
            "52.350001",
            "5.620001",
            "Een bestaande stadswandeling.",
            "Een volledige beschrijving van een bestaande stadswandeling.",
            "https://example.com/bestaande-stadswandeling-nieuwe-bron",
            "",
            "",
            "outdoor",
            "sun",
            "2026-05-13",
        ])
        sheet.append([
            "harderwijk",
            "",
            "",
            "activity",
            "Cultuur",
            "3 januari 2026",
            "Binnenstad",
            "Markt 3",
            "52.350002",
            "5.620002",
            "Deze rij mist een titel.",
            "Deze rij heeft expres een ontbrekende titel en mag niet importeren.",
            "https://example.com/missing-title",
            "",
            "",
            "outdoor",
            "sun",
            "2026-05-13",
        ])

        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        temp_file.close()
        workbook.save(temp_file.name)
        return temp_file.name

    def test_city_content_dry_run_does_not_write_database(self):
        path = self._create_city_content_excel_file()
        before_counts = {
            "events": Event.objects.count(),
            "venues": Venue.objects.count(),
            "tags": Tag.objects.count(),
        }
        stdout = StringIO()

        try:
            call_command(
                "import_city_content",
                file=path,
                city="harderwijk",
                type="outings",
                dry_run=True,
                no_report_file=True,
                stdout=stdout,
            )
        finally:
            if os.path.exists(path):
                os.unlink(path)

        self.assertEqual(Event.objects.count(), before_counts["events"])
        self.assertEqual(Venue.objects.count(), before_counts["venues"])
        self.assertEqual(Tag.objects.count(), before_counts["tags"])
        self.assertIn("DRY RUN ONLY: no database changes were written.", stdout.getvalue())

    def test_city_content_commit_imports_valid_rows_and_skips_errors_and_duplicates(self):
        existing = Event.objects.create(
            title="Bestaande stadswandeling",
            slug="bestaande-stadswandeling",
            kind=Event.Kind.ACTIVITY,
            city=self.city,
            summary="Niet overschrijven",
            source_url="https://example.com/originele-bron",
        )
        path = self._create_city_content_excel_file()
        report_dir = tempfile.mkdtemp()
        report_path = os.path.join(report_dir, "city_content_commit.json")
        stdout = StringIO()

        try:
            call_command(
                "import_city_content",
                file=path,
                city="harderwijk",
                type="outings",
                commit=True,
                report_file=report_path,
                stdout=stdout,
            )

            self.assertEqual(Event.objects.filter(city=self.city).count(), 2)
            imported = Event.objects.get(slug="nieuwe-stadswandeling")
            self.assertEqual(imported.kind, Event.Kind.ACTIVITY)
            self.assertEqual(imported.source, "city_content:outings")
            self.assertEqual(imported.category, self.category)

            existing.refresh_from_db()
            self.assertEqual(existing.summary, "Niet overschrijven")
            self.assertEqual(existing.source_url, "https://example.com/originele-bron")

            with open(report_path, encoding="utf-8") as report_file:
                report = json.load(report_file)

            self.assertEqual(report["summary"]["mode"], "commit")
            self.assertEqual(report["summary"]["total_rows"], 3)
            self.assertEqual(report["summary"]["valid_rows"], 2)
            self.assertEqual(report["summary"]["rows_with_errors"], 1)
            self.assertEqual(report["summary"]["imported_rows"], 1)
            self.assertEqual(report["summary"]["skipped_error_rows"], 1)
            self.assertEqual(report["summary"]["skipped_duplicate_rows"], 1)
            self.assertEqual(report["summary"]["not_imported_rows"], 2)

            actions = {row["row_number"]: row["action"] for row in report["rows"]}
            self.assertEqual(actions[2], "imported")
            self.assertEqual(actions[3], "skipped_duplicate")
            self.assertEqual(actions[4], "skipped_error")
            self.assertIn("COMMIT complete", stdout.getvalue())
        finally:
            if os.path.exists(path):
                os.unlink(path)
